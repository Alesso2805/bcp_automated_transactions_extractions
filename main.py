import re
import time
from datetime import datetime, timedelta
import openpyxl
import pandas as pd
import pyautogui
import pygetwindow as gw
import pyperclip
import win32con
import win32gui
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
import enter_telecredito

# We log to telecredito and open the window
# enter_telecredito.login_to_telecredito()

two_yesterday = (datetime.now() - timedelta(days=6)).strftime("%d%m%Y")
yesterday = (datetime.now() - timedelta(days=5)).strftime("%d%m%Y")

excel_path = rf'C:\Users\Flip\Desktop\Caja - {two_yesterday} - Agresivo.xlsx'
second_excel_path = rf'C:\Users\Flip\Desktop\Solicitudes - {two_yesterday} - AGR.xlsx'

df = pd.read_excel(excel_path, sheet_name='AccountDetail')
df_second = pd.read_excel(second_excel_path, sheet_name='Sheet1')

target_column = 'Unnamed: 11'
start_row = 8

screen_width, screen_height = pyautogui.size()
print(f"Resolución de pantalla detectada: {screen_width}x{screen_height}")

def buscar_y_extraer_valores(valor, monto, numero_operacion, df):
    print(f"Buscando valores para Dni={valor}, Cantidad={monto}, Nro Operacion={numero_operacion} y Procesado='NO'")
    fila = df[df['Numero Operacion Transferencia'] == numero_operacion]
    if not fila.empty:
        index = fila.index[0]
        nombre = fila['Nombre'].values[0]
        codigo_flip = fila['Codigo Flip'].values[0]
        cantidad = fila['Cantidad'].values[0]

        # Mark the row as processed
        df.at[index, 'Procesado'] = 'SI'

        # Save the changes to the Excel file

        workbook = load_workbook(second_excel_path)
        sheet = workbook['Sheet1']
        sheet.cell(row=index + 2, column=df.columns.get_loc('Procesado') + 1).value = 'SI'
        workbook.save(second_excel_path)

        print(f"Valores extraídos: Nombre={nombre}, Codigo Flip={codigo_flip}, Cantidad={cantidad}")
        return nombre, codigo_flip, cantidad
    print(f"No se encontraron valores para Dni={valor}, Cantidad={monto}, Nro Operacion={numero_operacion} y Procesado='NO'")
    return None, None, None

def focus_existing_window(title):
    windows = gw.getWindowsWithTitle(title)
    if windows:
        win = windows[0]
        win32gui.ShowWindow(win._hWnd, win32con.SW_MAXIMIZE)
        win32gui.SetForegroundWindow(win._hWnd)
        print(f"Ventana '{title}' enfocada.")
        return True
    return False

def wait_for_color(x, y, target_color, timeout=2):
    start_time = time.time()
    while time.time() - start_time < timeout:
        if pyautogui.pixel(x, y) == target_color:
            print(f"Color detectado en posición ({x}, {y}): {target_color}")
            return True
        time.sleep(0.5)
    print("Tiempo de espera agotado. No se detectó el color.")
    return False

def extraer_datos(texto: str):
    # Buscar número de operación
    match_operacion = re.search(r'Número de operación\s*\n\s*(\d+)', texto)
    numero_operacion = match_operacion.group(1) if match_operacion else None

    # Buscar monto
    match_monto = re.search(r'Monto\s*\n\s*\$ ([\d,]+\.\d{2})', texto)
    monto = match_monto.group(1) if match_monto else None

    # Buscar DNI
    match_dni = re.search(r'Nro documento\s*\n\s*(\d{8})', texto)
    dni = match_dni.group(1) if match_dni else None

    # Si falta alguno de los datos, devolver None
    if not (numero_operacion and monto and dni):
        return None

    return {
        "numero_operacion": numero_operacion,
        "monto": monto,
        "dni": dni
    }

def perform_action_and_insert_value(x, y, target_color, current_row):
    print(f"Color {target_color} detected at position ({x}, {y})")
    pyautogui.click(int(screen_width * 0.94), y)
    wait_for_color(int(screen_width * 0.746), int(screen_height * 0.412), (96,108,127))
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.hotkey('ctrl', 'c')
    copied_value = pyperclip.paste()
    print(f"Copied value: {copied_value}")

    # Extract data using the extraer_datos function
    datos = extraer_datos(copied_value)
    if datos:
        numero_operacion = datos["numero_operacion"]
        monto = datos["monto"]
        dni = datos["dni"]
        print(f"Numero de Operacion: {numero_operacion}, Monto: {monto}, DNI: {dni}")
    else:
        print("No se pudieron extraer todos los datos necesarios.")
        return

    while current_row < (len(df) + 2) and pd.notna(df.at[current_row, target_column]):
        current_row += 1
    df.at[current_row, target_column] = dni
    print(f"Value '{dni}' inserted into row {current_row}, column {target_column}")
    workbook = load_workbook(excel_path)
    sheet = workbook['AccountDetail']
    sheet.cell(row=current_row + 1, column=df.columns.get_loc(target_column) + 1).value = dni
    print(f"Obteniendo valores del segundo Excel para Dni={dni}, Nro Operacion={numero_operacion} y Monto={monto}")
    nombre, codigo_flip, cantidad = buscar_y_extraer_valores(dni, monto, numero_operacion, df_second)
    if nombre and codigo_flip and cantidad:
        print("La fila actual es: ", current_row)
        df.at[current_row, 'Unnamed: 9'] = nombre
        df.at[current_row, 'Unnamed: 15'] = codigo_flip
        df.at[current_row, 'Unnamed: 13'] = cantidad
        sheet.cell(row=current_row + 1, column=df.columns.get_loc('Unnamed: 9') + 1).value = nombre
        sheet.cell(row=current_row + 1, column=df.columns.get_loc('Unnamed: 15') + 1).value = codigo_flip
        sheet.cell(row=current_row + 1, column=df.columns.get_loc('Unnamed: 13') + 1).value = cantidad
        print(f"Nombre '{nombre}', Codigo Flip '{codigo_flip}' y Cantidad '{cantidad}' insertados en la fila {current_row}")
    else:
        columns_to_check = ['Unnamed: 9', 'Unnamed: 15', 'Unnamed: 13']
        for col in columns_to_check:
            df.at[current_row, col] = 'PENDIENTE'
            sheet.cell(row=current_row + 1, column=df.columns.get_loc(col) + 1).value = 'PENDIENTE'
        current_row += 1
        print(f"Valores no encontrados, se insertó 'PENDIENTE' en las columnas designadas para la fila {current_row}")

    workbook.save(excel_path)
    print("DataFrame saved to Excel file")
    pyautogui.click(int(screen_width * 0.21), int(screen_height * 0.289))
    time.sleep(0.5)

def search_color_on_screen(target_color, current_row, timeout=3):
    start_time = time.time()
    while time.time() - start_time < timeout:
        screenshot = pyautogui.screenshot()
        width, height = screenshot.size
        for x in range(width):
            for y in range(height):
                if screenshot.getpixel((x, y)) == target_color:
                    perform_action_and_insert_value(x, y, target_color, current_row)
                    return (x, y)
        time.sleep(0.5)
    print(f"Timeout reached. Color {target_color} not detected on the screen.")
    return None

# Variable global para almacenar la coordenada base del primer botón
base_x = None

def go_to_next_page(current_page, num_pages):
    global base_x  # Usamos la variable global para mantener la posición del primer botón
    y = 881  # Coordenada Y fija
    target_color = (255, 120, 0)  # Color objetivo

    # Si no se ha detectado aún, buscamos el primer botón y guardamos su posición
    if base_x is None:
        print("Buscando el botón de la página 1...")
        screen_width, screen_height = pyautogui.size()  # Obtener el tamaño de la pantalla
        for x in range(int(screen_width * 0.49), screen_width):
            try:
                if pyautogui.pixelMatchesColor(x, y, target_color, tolerance=10):
                    base_x = x  # Guardamos la posición del primer botón
                    pyautogui.moveTo(base_x, y)
                    print(f"Botón de la página 1 encontrado en ({base_x}, {y})")
                    break
            except Exception as e:
                print(f"Error al obtener el color en ({x}, {y}): {e}")
                continue
        else:
            print("No se encontró el botón de la página 1.")
            return None, None

    # Calcular la posición del botón de la siguiente página basándose en base_x
    spacing = 58  # Ajusta este valor si es necesario
    x = base_x + (current_page * spacing)

    if pyautogui.onScreen(x, y):
        pyautogui.moveTo(x, y)
        pyautogui.click(x, y)
        print(f"Haciendo clic en la página {current_page + 1} en ({x}, {y})")
        time.sleep(1)
    else:
        print(f"El botón de la página {current_page + 1} no está en la pantalla.")

    return base_x, y

def main():
    if focus_existing_window("Nuevo Telecrédito"):
        print("La ventana ya estaba abierta y fue enfocada.")
    time.sleep(0.5)
    pyautogui.click(int(screen_width * 0.50), int(screen_height * 0.15))
    time.sleep(0.5)
    pyautogui.hotkey('tab')
    pyautogui.typewrite('agr')
    time.sleep(0.4)
    pyautogui.click(int(screen_width * 0.50), int(screen_height * 0.24))
    wait_for_color(int(screen_width * 0.39), int(screen_height * 0.74), (232, 246, 252))
    pyautogui.scroll(-500)
    time.sleep(0.1)
    pyautogui.click(int(screen_width * 0.25), int(screen_height * 0.60))
    pyautogui.click(int(screen_width * 0.25), int(screen_height * 0.30))
    wait_for_color(int(screen_width * 0.316), int(screen_height * 0.804), (212, 102, 40))
    pyautogui.click(int(screen_width * 0.35), int(screen_height * 0.36))
    time.sleep(0.5)
    for i in range(4):
        pyautogui.hotkey('tab')
        pyautogui.hotkey('backspace')
    pyautogui.click(int(screen_width * 0.22), int(screen_height * 0.72))
    pyautogui.typewrite(two_yesterday)
    pyautogui.hotkey('tab')
    pyautogui.typewrite(yesterday)
    pyautogui.hotkey('tab')
    pyautogui.hotkey('enter')
    wait_for_color(int(screen_width * 0.823), int(screen_height * 0.878), (96,108,127))
    pyautogui.scroll(-4000)
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.hotkey('ctrl', 'c')
    copied_text = pyperclip.paste()
    page_numbers = re.findall(r'\d+', copied_text)
    if page_numbers:
        num_pages = int(page_numbers[-1]) + 1
        print(f"Number of pages detected: {num_pages}")
    else:
        print("No page numbers found in the copied text.")

    current_row = start_row
    base_x, y = go_to_next_page(0, num_pages)  # Get the coordinates for page 1
    if base_x is None or y is None:
        return
    for index, row in df.iloc[current_row - 1:].iterrows():
        numero_operacion = row['Unnamed: 5']
        monto = row['Unnamed: 3']  # Ensure the correct row is accessed
        if monto < 0:
            print(f"Skipping negative monto: {monto}")
            print(f"Skipping numero de operacion {numero_operacion}")
            current_row += 1
            print(f"Fila actualizada al comenzar: {current_row}")
            continue
        match_found = False
        while not match_found:
            for page in range(num_pages):
                time.sleep(0.5)
                pyautogui.hotkey('ctrl', 'f')
                pyautogui.typewrite(str(numero_operacion))
                pyautogui.press('enter')
                pyautogui.press('enter')
                pyautogui.scroll(40)
                time.sleep(0.5)
                if search_color_on_screen((255, 150, 50), current_row):
                    print(f"Match found for {numero_operacion} on page {page + 1}")
                    current_row += 1
                    print(f"Fila actualizada al terminar: {current_row}")
                    match_found = True
                    break
                else:
                    pyautogui.scroll(-4000)
                    time.sleep(0.5)
                    go_to_next_page(page, num_pages)
                    time.sleep(0.5)
                    print(f"No match found for {numero_operacion} on page {page + 1}, moving to next page")
            if not match_found:
                print(f"Value {numero_operacion} not found, moving to the next cell")
                pyautogui.moveTo(base_x, y)  # Click on the coordinates of page 1
                pyautogui.click(base_x, y)
                current_row += 1
                break

    # Filter rows with 'PENDIENTE' and save to a new sheet
    pendientes_df = df[df.isin(['PENDIENTE']).any(axis=1)]
    custom_headers = ['Fecha', 'Fecha Valuta', 'Descripción Operación', 'Monto', 'Sucursal - Agencia', 'N° Operación',
                      'Usuario', 'x', 'x', 'NOMBRE', 'x', 'DNI', 'x', 'Monto', 'x', 'CODIGO FLIP']
    pendientes_df.columns = custom_headers

    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
        pendientes_df.to_excel(writer, sheet_name='PENDIENTES', index=False)

    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook['PENDIENTES']

    # Center align all cells and adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Save the workbook
    workbook.save(excel_path)
    print("Rows with 'PENDIENTE' moved to a new sheet named 'PENDIENTES' with custom headers and formatted cells")

main()